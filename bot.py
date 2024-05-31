import discord
from discord.ext import commands, tasks
from pymongo import MongoClient
from datetime import datetime, timedelta
import pytz
from bson.objectid import ObjectId
from dotenv import load_dotenv
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
load_dotenv()

# MongoDB setup
MONGO_URI = os.getenv("MONGO_URI")
client = MongoClient(MONGO_URI)
db = client["working_hours_db"]

# Bot setup
intents = discord.Intents.default()
bot = commands.Bot(command_prefix='/', intents=intents)

# Define JST timezone
JST = pytz.timezone('Asia/Tokyo')

def get_collection(guild_id):
    return db[str(guild_id)]

def calculate_total_minutes(entries):
    total_seconds = sum((entry['end_time'] - entry['start_time']).total_seconds() for entry in entries if entry['end_time'])
    return total_seconds / 60

def generate_unique_id(collection, date):
    date_str = date.strftime('%y%m%d')
    count = collection.count_documents({"unique_id": {"$regex": f"^{date_str}-"}}) + 1
    return f"{date_str}-{count:03d}"

def generate_excel(filename, headers, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    for col_num in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 20

    workbook.save(filename)

@bot.event
async def on_ready():
    print(f'Bot is ready. Logged in as {bot.user}')
    check_work_sessions.start()

@bot.slash_command(name="shukkin", description="出勤する")
async def start_work(ctx):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    user_id = ctx.author.id
    discord_name = str(ctx.author)
    start_time = datetime.now(JST)
    unique_id = generate_unique_id(collection, start_time)

    # Check for an existing active work session
    existing_entry = collection.find_one({"user_id": user_id, "end_time": None})
    if existing_entry:
        await ctx.respond("You already have an active work session.")
        return

    # Create a new entry in MongoDB
    entry = {
        "user_id": user_id,
        "discord_name": discord_name,
        "start_time": start_time,
        "end_time": None,
        "unique_id": unique_id
    }
    collection.insert_one(entry)

    embed = discord.Embed(title="出勤を開始しました", description=f"{ctx.user.mention} さんおかえりなさい！\n {start_time.strftime('%Y-%m-%d %H:%M')} に勤務を開始しました", color=discord.Color.green())
    embed.add_field(name="勤務データID", value=unique_id)
    embed.set_footer(text="Powered by NickyBoy", icon_url="https://i.imgur.com/QfmDKS6.png")
    await ctx.respond(embed=embed)

@bot.slash_command(name="taikin", description="退勤する")
async def end_work(ctx):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    user_id = ctx.author.id
    end_time = datetime.now(JST)

    # Find the last entry without an end time
    entry = collection.find_one({"user_id": user_id, "end_time": None})
    if not entry:
        await ctx.respond("No work session to end.")
        return

    collection.update_one({"_id": entry["_id"]}, {"$set": {"end_time": end_time}})

    embed = discord.Embed(title="退勤しました", description=f"{ctx.author.mention}さんお疲れさまでした!\n {end_time.strftime('%Y-%m-%d %H:%M')} に退勤しました", color=discord.Color.red())
    embed.set_footer(text="Powered by NickyBoy", icon_url="https://i.imgur.com/QfmDKS6.png")
    await ctx.respond(embed=embed)

@tasks.loop(minutes=10)
async def check_work_sessions():
    now = datetime.now(JST)
    threshold = now - timedelta(hours=10)

    for guild in bot.guilds:
        collection = get_collection(guild.id)
        active_entries = collection.find({"end_time": None, "start_time": {"$lte": threshold}})
        for entry in active_entries:
            user = guild.get_member(entry["user_id"])
            if user:
                await user.send(f"Your work session started at {entry['start_time'].strftime('%Y-%m-%d %H:%M')} has exceeded 10 hours. Please remember to end it using /taikin.")

@bot.slash_command(name="shuusei", description="出勤時間・退勤時間を修正")
async def edit_work(ctx, unique_id: discord.Option(str, "勤務データIDを指定", required=True), new_start: discord.Option(str, "新しい勤務開始時間 例: 2024-01-01 0:00", required=True), new_end: discord.Option(str, "新しい退勤時間 例: 2024-01-02 0:00", required=True)):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    
    entry = collection.find_one({"unique_id": unique_id})
    if not entry:
        await ctx.respond("そのIDは存在しません")
        return
    
    if entry['user_id'] != ctx.author.id:
        await ctx.respond("他人の出勤データは編集できません", ephemeral=True)
        return

    try:
        new_start_time = datetime.strptime(new_start, '%Y-%m-%d %H:%M').replace(tzinfo=JST)
        new_end_time = datetime.strptime(new_end, '%Y-%m-%d %H:%M').replace(tzinfo=JST)
    except ValueError:
        await ctx.respond("日付と時間の形式が違います 例: 2024-01-01 0:00")
        return

    old_start_time = entry['start_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M')
    old_end_time = entry['end_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M') if entry['end_time'] else "Ongoing"

    collection.update_one({"unique_id": unique_id}, {"$set": {"start_time": new_start_time, "end_time": new_end_time}})

    embed = discord.Embed(title="出勤データを修正しました", description=f"ID: {unique_id} のデータを修正しました", color=discord.Color.blue())
    embed.add_field(name="スタッフ", value=f"{entry['discord_name']}", inline=False)
    embed.add_field(name="修正前の出勤時間", value=old_start_time, inline=True)
    embed.add_field(name="修正前の退勤時間", value=old_end_time, inline=True)
    embed.add_field(name="修正後の出勤時間", value=new_start_time.strftime('%Y-%m-%d %H:%M'), inline=True)
    embed.add_field(name="修正後の退勤時間", value=new_end_time.strftime('%Y-%m-%d %H:%M'), inline=True)
    embed.set_footer(text="Powered by NickyBoy", icon_url="https://i.imgur.com/QfmDKS6.png")

    await ctx.respond(embed=embed)

@bot.slash_command(name="admintaikin", description="強制的に退勤させる")
@commands.has_permissions(administrator=True)
async def admin_end_work(ctx, unique_id: discord.Option(str, "勤務データIDを指定", required=True)):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    end_time = datetime.now(JST)

    entry = collection.find_one({"unique_id": unique_id, "end_time": None})
    if not entry:
        await ctx.respond("No active work session found with the given ID.")
        return

    collection.update_one({"unique_id": unique_id}, {"$set": {"end_time": end_time}})

    embed = discord.Embed(title="強制退勤しました", description=f"ID: {unique_id} の勤務データが強制的に退勤されました", color=discord.Color.red())
    embed.set_footer(text="Powered by NickyBoy", icon_url="https://i.imgur.com/QfmDKS6.png")
    await ctx.respond(embed=embed)

@bot.slash_command(name="adminshuusei", description="強制的に出勤時間・退勤時間を修正")
@commands.has_permissions(administrator=True)
async def admin_edit_work(ctx, unique_id: discord.Option(str, "勤務データIDを指定", required=True), new_start: discord.Option(str, "新しい勤務開始時間 例: 2024-01-01 0:00", required=True), new_end: discord.Option(str, "新しい退勤時間 例: 2024-01-02 0:00", required=True)):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    
    entry = collection.find_one({"unique_id": unique_id})
    if not entry:
        await ctx.respond("そのIDは存在しません")
        return

    try:
        new_start_time = datetime.strptime(new_start, '%Y-%m-%d %H:%M').replace(tzinfo=JST)
        new_end_time = datetime.strptime(new_end, '%Y-%m-%d %H:%M').replace(tzinfo=JST)
    except ValueError:
        await ctx.respond("日付と時間の形式が違います 例: 2024-01-01 0:00")
        return

    old_start_time = entry['start_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M')
    old_end_time = entry['end_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M') if entry['end_time'] else "Ongoing"

    collection.update_one({"unique_id": unique_id}, {"$set": {"start_time": new_start_time, "end_time": new_end_time}})

    embed = discord.Embed(title="強制的に出勤データを修正しました", description=f"ID: {unique_id} のデータを修正しました", color=discord.Color.blue())
    embed.add_field(name="スタッフ", value=f"{entry['discord_name']}", inline=False)
    embed.add_field(name="修正前の出勤時間", value=old_start_time, inline=True)
    embed.add_field(name="修正前の退勤時間", value=old_end_time, inline=True)
    embed.add_field(name="修正後の出勤時間", value=new_start_time.strftime('%Y-%m-%d %H:%M'), inline=True)
    embed.add_field(name="修正後の退勤時間", value=new_end_time.strftime('%Y-%m-%d %H:%M'), inline=True)
    embed.set_footer(text="Powered by NickyBoy", icon_url="https://i.imgur.com/QfmDKS6.png")

    await ctx.respond(embed=embed)

@bot.slash_command(name="kakunin", description="出勤データを確認")
async def check_work(ctx):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    user_id = ctx.author.id

    entries = list(collection.find({"user_id": user_id}).sort("start_time", -1).limit(10))
    total_minutes = calculate_total_minutes(entries)

    embed = discord.Embed(title="最近の出勤データ (直近１０個)", color=discord.Color.gold())
    for entry in entries:
        start = entry['start_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M')
        end = entry['end_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M') if entry['end_time'] else "Ongoing"
        embed.add_field(name=f"勤務データID: {entry['unique_id']}", value=f"Start: {start}\nEnd: {end}", inline=False)
    embed.add_field(name="合計出勤時間(分)", value=f"{total_minutes:.2f}", inline=False)
    await ctx.respond(embed=embed, ephemeral=True)

@bot.slash_command(name="list", description="日付の範囲を指定して、期間内の従業員の出勤時間を算出")
async def list_work(ctx, start_date: discord.Option(str, "日付の範囲指定 例:2024-01-01", required=True), end_date: discord.Option(str, "日付の範囲指定 例:2024-01-01", required=True)):
    guild_id = ctx.guild.id
    collection = get_collection(guild_id)
    try:
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=JST)
            end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=JST)
            query = {"end_time": {"$gte": start_date, "$lt": end_date + timedelta(days=1)}}
        else:
            query = {}
    except ValueError:
        await ctx.respond("日付と時間の形式が違います 例: 2024-01-01")
        return

    entries = list(collection.find(query))

    user_minutes = {}
    for entry in entries:
        if entry['end_time']:
            user_minutes.setdefault(entry['discord_name'], 0)
            user_minutes[entry['discord_name']] += (entry['end_time'] - entry['start_time']).total_seconds()

    embed = discord.Embed(title="各従業員の出勤時間(分)", color=discord.Color.purple())
    for user, total_seconds in user_minutes.items():
        total_minutes = total_seconds / 60
        embed.add_field(name=user, value=f"合計出勤時間(分): {total_minutes:.2f}", inline=False)

    await ctx.respond(embed=embed, ephemeral=True)

@bot.slash_command(name="exportdata", description="個別の出勤データをエクセルファイルに出力")
async def export_data(ctx, start_date: discord.Option(str, "日付の範囲指定 例:2024-01-01", required=True), end_date: discord.Option(str, "日付の範囲指定 例:2024-01-02", required=True)):
    await ctx.respond("エクセルファイルを準備しています...")

    guild_id = ctx.guild.id
    collection = get_collection(guild_id)

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=JST)
        end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=JST)
        query = {"end_time": {"$gte": start_date, "$lt": end_date + timedelta(days=1)}}
    except ValueError:
        await ctx.respond("日付と時間の形式が違います 例: 2024-01-01")
        return

    entries = list(collection.find(query))

    data = []
    for entry in entries:
        if entry['end_time']:
            start_time = entry['start_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M')
            end_time = entry['end_time'].astimezone(JST).strftime('%Y-%m-%d %H:%M')
            total_minutes = (entry['end_time'] - entry['start_time']).total_seconds() / 60
            data.append([entry['discord_name'], start_time, end_time, total_minutes])

    headers = ["Employee Name", "Start Time", "End Time", "Total Minutes"]
    filename = f"work_data_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    generate_excel(filename, headers, data)

    await ctx.respond(file=discord.File(filename))
    os.remove(filename)  # Delete the file after sending

@bot.slash_command(name="exporttotal", description="各従業員の総出勤時間(分)をエクセルファイルに出力")
async def export_total(ctx, start_date: discord.Option(str, "日付と時刻の範囲指定 例:2024-01-01", required=True), end_date: discord.Option(str, "日付と時刻の範囲指定 例:2024-01-01", required=True)):
    await ctx.respond("エクセルファイルを準備しています...", ephemeral=True)

    guild_id = ctx.guild.id
    collection = get_collection(guild_id)

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=JST)
        end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=JST)
        query = {"end_time": {"$gte": start_date, "$lt": end_date + timedelta(days=1)}}
    except ValueError:
        await ctx.respond("日付の形式が違います 例: 2024-01-01")
        return

    entries = list(collection.find(query))

    user_minutes = {}
    for entry in entries:
        if entry['end_time']:
            user_minutes.setdefault(entry['discord_name'], 0)
            user_minutes[entry['discord_name']] += (entry['end_time'] - entry['start_time']).total_seconds()

    data = [[user, total_seconds / 60] for user, total_seconds in user_minutes.items()]
    headers = ["Employee Name", "Total Minutes"]
    filename = f"total_hours_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    generate_excel(filename, headers, data)

    await ctx.respond(file=discord.File(filename))
    os.remove(filename)  # Delete the file after sending

bot.run(os.getenv("DISCORD_BOT_TOKEN"))